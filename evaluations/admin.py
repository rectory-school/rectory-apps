import logging
from typing import Optional
from tempfile import NamedTemporaryFile
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from django.contrib import admin, messages
from django.http import HttpRequest, HttpResponse
from django.urls import URLPattern, path
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
from django.db.models import Q, Count
from django.utils import timezone


from adminsortable2.admin import SortableInlineAdminMixin

from simple_history.models import HistoricalRecords

from blackbaud.models import Student

from . import models

log = logging.getLogger(__name__)


@admin.register(models.QuestionSet)
class QuestionSetAdmin(admin.ModelAdmin):
    """Admin for a question set"""

    search_fields = ["name"]


@admin.register(models.FreeformQuestion)
class FreeformQuestionAdmin(admin.ModelAdmin):
    """Admin for a freeform question"""

    list_display = ["__str__", "question_set"]
    autocomplete_fields = ["question_set"]
    list_filter = ["question_set"]

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("question_set")


class MultipleChoiceAnswerInline(admin.TabularInline):
    """Inline for a multiple choice question"""

    model = models.MultipleChoiceAnswer
    extra = 0


@admin.register(models.MultipleChoiceQuestion)
class MultipleChoiceQuestionAdmin(admin.ModelAdmin):
    """Admin for a multiple choice question"""

    list_display = ["__str__", "question_set"]
    autocomplete_fields = ["question_set"]
    list_filter = ["question_set"]
    inlines = (MultipleChoiceAnswerInline,)

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("question_set")


class TagInline(admin.TabularInline):
    """Inline for tags in categories"""

    model = models.Tag
    extra = 0


@admin.register(models.TagCategory)
class TagCategoryAdmin(admin.ModelAdmin):
    """Admin for a tag category"""

    inlines = [TagInline]
    readonly_fields = ["system_managed"]
    list_display = [
        "__str__",
        "system_managed",
        "show_as_filter",
    ]

    def has_change_permission(
        self, request, obj: Optional[models.TagCategory] = None
    ) -> bool:
        if obj and obj.system_managed:
            return False

        return super().has_change_permission(request, obj)

    def has_delete_permission(
        self, request, obj: Optional[models.TagCategory] = None
    ) -> bool:
        if obj and obj.system_managed:
            return False

        return super().has_change_permission(request, obj)


@admin.register(models.Tag)
class TagAdmin(admin.ModelAdmin):
    search_fields = ["category__name", "name"]
    list_display = ["__str__", "category", "created_at"]
    list_filter = ["category"]

    def has_add_permission(self, *args, **kwargs) -> bool:
        return False

    def has_change_permission(self, *args, **kwargs) -> bool:
        return False

    def has_delete_permission(self, *args, **kwargs) -> bool:
        return False


@admin.register(models.Evaluation)
class EvaluationAdmin(admin.ModelAdmin):
    """Admin for individual evaluations"""

    autocomplete_fields = ["tags", "student", "question_set"]
    search_fields = ["student__given_name", "student__family_name"]
    list_display = ["__str__", "student", "completed_at"]

    list_filter = (
        ("question_set", admin.RelatedOnlyFieldListFilter),
        "created_at",
        "completed_at",
    )

    actions = [
        "send_reminder_emails",
    ]

    history = HistoricalRecords()

    def get_list_filter(self, request: HttpRequest):
        base_filters = list(super().get_list_filter(request))

        for category in models.TagCategory.objects.filter(show_as_filter=True):
            base_filters.append(get_filter_for_tag_category_selection(category))

        return base_filters

    @admin.action(description="Send reminder emails")
    def send_reminder_emails(self, request, objs):
        """Send a reminder email to the students with incomplete evaluations"""

        raise NotImplementedError()

    def changelist_view(
        self,
        request: HttpRequest,
        extra_context: Optional[dict] = None,
    ):
        extra_context = extra_context or {}

        changelist = self.get_changelist_instance(request)
        qs = changelist.get_queryset(request)
        q_starting_available = Q(available_starting_at__isnull=True)
        q_starting_available |= Q(available_starting_at__gte=timezone.now())

        q_available_until = Q(available_until__isnull=True)
        q_available_until |= Q(available_until__lte=timezone.now())
        q_complete = Q(completed_at__isnull=False)

        qs_available = qs.filter(q_starting_available & q_available_until)
        qs_completed_available = qs_available.filter(q_complete)

        qs_unavailable = qs.filter(~q_starting_available | ~q_available_until)
        qs_unavailable_complete = qs_unavailable.filter(q_complete)

        available = qs_available.count()
        available_complete = qs_completed_available.count()

        unavailable = qs_unavailable.count()
        unavailable_complete = qs_unavailable_complete.count()

        extra_context["available"] = available
        extra_context["available_complete"] = available_complete
        if available:
            extra_context["available_complete_pct"] = (
                available_complete / available * 100
            )

        extra_context["unavailable"] = unavailable
        extra_context["unavailable_complete"] = unavailable_complete
        if unavailable:
            extra_context["unavailable_complete_pct"] = (
                unavailable_complete / unavailable
            )

        return super().changelist_view(request, extra_context)

    def clear_answer_view(self, request, pk: int):
        """Do the needful"""

        with transaction.atomic():
            evaluation = get_object_or_404(models.Evaluation, pk=pk)
            models.EvaluationMultipleChoiceAnswer.objects.filter(
                evaluation=evaluation
            ).delete()
            models.EvaluationFreeformAnswer.objects.filter(
                evaluation=evaluation
            ).delete()
            evaluation.completed_at = None
            evaluation.save()

        messages.success(request, "Selected evaluation answers have been cleared")
        return redirect("admin:evaluations_evaluation_change", evaluation.pk)

    def get_urls(self) -> list[URLPattern]:
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:pk>/clear/",
                self.admin_site.admin_view(self.clear_answer_view),
                name="evaluations_clear_answers",
            )
        ]

        return custom_urls + urls


class ReportCategorizerCategoryInline(SortableInlineAdminMixin, admin.TabularInline):
    model = models.ReportCategorizerTagCategory
    extra = 0


@admin.register(models.ReportCategorizer)
class ReportCategorizerAdmin(admin.ModelAdmin):
    """Admin for report categorizer"""

    inlines = [ReportCategorizerCategoryInline]


class UploadConfigurationTagInline(admin.TabularInline):
    model = models.UploadConfigurationTagCategory
    extra = 0


@admin.register(models.UploadConfiguration)
class UploadConfigurationAdmin(admin.ModelAdmin):
    inlines = [UploadConfigurationTagInline]

    def get_urls(self) -> list[URLPattern]:
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:pk>/download",
                self.admin_site.admin_view(self.download_template),
                name="evaluations_uploadconfiguration_download_template",
            ),
            path(
                "<int:pk>/upload",
                self.admin_site.admin_view(self.upload_template),
                name="evaluations_uploadconfiguration_upload_template",
            ),
        ]

        return urls + custom_urls

    @transaction.atomic
    def upload_template(self, request: HttpRequest, pk: int) -> HttpResponse:
        upload_obj = get_object_or_404(models.UploadConfiguration, pk=pk)
        excel_data = load_workbook(
            request.FILES["file"],
            data_only=True,
            read_only=True,
        )
        ws = excel_data.active

        rows = ws.rows
        header_row = next(rows)

        headers_by_index = {cell.value: i for i, cell in enumerate(header_row)}
        expected_headers = {"Student Email", "Evaluation Title"}
        tag_category_by_column: dict[models.TagCategory, int] = {}

        for tag_category in upload_obj.tag_categories.all():
            expected_headers.add(tag_category.category.name)
            column = headers_by_index.get(tag_category.category.name, None)
            if column is not None:
                tag_category_by_column[tag_category.category] = column

        extra_headers = headers_by_index.keys() - expected_headers
        missing_headers = expected_headers - headers_by_index.keys()

        for header in extra_headers:
            messages.warning(
                request,
                f'The header "{header}" was not configured and thus ignored',
            )

        for header in missing_headers:
            messages.error(
                request,
                f'The header "{header}" was missing',
            )

        if missing_headers:
            # Bail early
            return redirect("admin:evaluations_uploadconfiguration_change", pk)

        # Grab all the students by email
        student_email_rows = ws.rows
        next(student_email_rows)  # Throw out the first row

        student_email_header_pos = headers_by_index["Student Email"]
        student_emails: set[str] = set()
        for row in student_email_rows:
            student_email = row[student_email_header_pos].value
            student_emails.add(student_email.strip())

        students_by_email = {
            obj.email: obj for obj in Student.objects.filter(email__in=student_emails)
        }

        tags_by_category: dict[models.TagCategory, dict[str, models.Tag]] = {}
        for category in tag_category_by_column:
            tags_by_category[category] = {}
            for tag in category.tags.all():
                tags_by_category[category][tag.value.lower()] = tag

        # Create the relevant tags
        tag_rows = ws.rows
        next(tag_rows)  # Skip the first row
        for row in tag_rows:
            for category, column in tag_category_by_column.items():
                row_value = row[column].value
                if not row_value:
                    continue
                normalized_value = str(row_value).strip().lower()
                if normalized_value not in tags_by_category[category]:
                    tag = models.Tag.objects.create(
                        category=category,
                        value=str(row_value).strip(),
                    )
                    tags_by_category[category][normalized_value] = tag
                    messages.info(request, f"Created tag {category.name}/{tag.value}")

        upload_id = uuid4()
        upload_tag = models.Tag.objects.create(
            category=models.TagCategory.objects.get(name="Upload ID"),
            value=str(upload_id),
        )

        # Finally - create the actual evaluations
        rows = ws.rows
        next(rows)

        for i, row in enumerate(rows):
            student_email = row[headers_by_index["Student Email"]].value
            try:
                student = students_by_email[student_email]
            except KeyError:
                messages.error(
                    request,
                    f"Student with email {student_email} could not be found",
                )
                continue

            title = row[headers_by_index["Evaluation Title"]].value
            if not title:
                messages.error(request, f"Evaluation title was missing on row {i+2}")
                continue

            tags: set[models.Tag] = set()

            for category in tags_by_category:
                value = row[tag_category_by_column[category]].value
                if value:
                    tags.add(tags_by_category[category][str(value).lower().strip()])

            evaluation = models.Evaluation.objects.create(
                name=title,
                student=student,
                question_set=upload_obj.question_set,
            )
            for tag in tags:
                evaluation.tags.add(tag)
            evaluation.tags.add(upload_tag)

        messages.success(request, f"Upload successful, upload ID is {upload_id}")
        return redirect("admin:evaluations_uploadconfiguration_change", pk)

    def download_template(self, request: HttpRequest, pk: int) -> HttpResponse:
        """Generate and download the template for uploading"""

        obj = get_object_or_404(models.UploadConfiguration, pk=pk)

        wb = Workbook()
        ws = wb.active
        ws.title = "Template"

        headers = ["Student Email", "Evaluation Title"]

        for tag_category in obj.tag_categories.order_by("category__name"):
            headers.append(tag_category.category.name)

        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        resp = HttpResponse(
            stream,
            headers={
                "Content-Type": (
                    "application/vnd."
                    "openxmlformats-officedocument."
                    "spreadsheetml."
                    "sheet"
                ),
                "Content-Disposition": f"attachment; filename={obj.name}.xlsx",
            },
        )

        return resp


def get_filter_for_tag_category_selection(category: models.TagCategory):
    """Get a single tag category filter"""

    class TagCategoryFilter(admin.SimpleListFilter):
        """A filter that will include a tag category
        with all the relevant tags available"""

        title = category.name
        parameter_name = f"tag_category_{category.pk}"

        def lookups(self, request, model_admin):
            qs = model_admin.get_queryset(request)

            # This breaks if other tag category filters are enabled
            non_tag_query_attrs = dict(
                [
                    (param, val)
                    for param, val in request.GET.items()
                    if not param.startswith("tag_category_")
                ]
            )

            qs = qs.filter(**non_tag_query_attrs)

            tag_query_attrs = dict(
                [
                    (param, val)
                    for param, val in request.GET.items()
                    if param.startswith("tag_category_")
                ]
            )

            for val in tag_query_attrs.values():
                tag = models.Tag.objects.get(pk=val)
                qs = qs.filter(tags=tag)

            # Simple return
            tags = models.Tag.objects.filter(category=category)
            tags = tags.filter(evaluation__in=qs)
            tags = tags.annotate(evaluation_count=Count("evaluation"))
            tags = tags.filter(evaluation_count__gt=0)
            return [(tag.pk, tag.value) for tag in tags]

        def queryset(self, request, queryset):
            value = self.value()

            if value:
                tag = models.Tag.objects.filter(category=category, pk=self.value())
                queryset = queryset.filter(tags__in=tag)

            return queryset

    return TagCategoryFilter
